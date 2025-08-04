#!/usr/bin/env python3
"""
Export Monte Carlo Simulation Results to Excel
Creates a comprehensive Excel workbook with multiple sheets, charts, and formatted tables
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import (
    BarChart, LineChart, PieChart, ScatterChart, 
    Reference, Series
)
from openpyxl.chart.axis import DateAxis
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
import json
from datetime import datetime
import warnings
warnings.filterwarnings('ignore')

# Import the simulation functions
from consumer_harm_monte_carlo import (
    run_monte_carlo_simulation, 
    calculate_statistics, 
    PARAMS, 
    ANNUAL_TRANSACTIONS,
    N_SIMULATIONS
)

class ExcelExporter:
    def __init__(self, filename="Consumer_Harm_Analysis.xlsx"):
        self.filename = filename
        self.wb = Workbook()
        self.wb.remove(self.wb.active)  # Remove default sheet
        
        # Define styles
        self.header_font = Font(bold=True, color="FFFFFF", size=12)
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        self.title_font = Font(bold=True, size=16)
        self.subtitle_font = Font(bold=True, size=14)
        self.section_font = Font(bold=True, size=12)
        
        self.currency_format = '$#,##0.00'
        self.percent_format = '0.0%'
        self.number_format = '#,##0'
        
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Color schemes
        self.good_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        self.bad_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        self.neutral_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    def create_summary_sheet(self, stats, scenario_results):
        """Create executive summary sheet"""
        ws = self.wb.create_sheet("Executive Summary")
        
        # Title
        ws['A1'] = "Consumer Harm Monte Carlo Analysis"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:F1')
        
        ws['A2'] = f"Analysis Date: {datetime.now().strftime('%B %d, %Y')}"
        ws['A3'] = f"Simulations: {N_SIMULATIONS:,} | Annual Transactions: {ANNUAL_TRANSACTIONS:,.0f}"
        ws.merge_cells('A2:F2')
        ws.merge_cells('A3:F3')
        
        # Key Findings Section
        ws['A5'] = "KEY FINDINGS"
        ws['A5'].font = self.subtitle_font
        ws.merge_cells('A5:F5')
        
        # Status Quo Statistics
        row = 7
        ws[f'A{row}'] = "Status Quo Analysis"
        ws[f'A{row}'].font = self.section_font
        row += 1
        
        key_stats = [
            ("Mean Consumer Harm", stats['Mean Harm'], self.currency_format),
            ("Median Consumer Harm", stats['Median Harm'], self.currency_format),
            ("95th Percentile Harm", stats['95th Percentile'], self.currency_format),
            ("Maximum Harm Observed", stats['Max Harm'], self.currency_format),
            ("", "", ""),  # Blank row
            ("Customers with Zero Harm", stats['Customers with Zero Harm'], self.number_format),
            ("Customers with Harm > $1,000", stats['Customers with Harm > $1000'], self.number_format),
            ("Customers with Harm > $5,000", stats['Customers with Harm > $5000'], self.number_format),
            ("", "", ""),  # Blank row
            ("Annual Industry Impact (Mean)", stats['Annual Industry Impact (Mean)'], self.currency_format),
            ("Annual Industry Impact (95th %ile)", stats['Annual Industry Impact (95th %ile)'], self.currency_format),
        ]
        
        for stat_name, stat_value, format_str in key_stats:
            if stat_name:  # Skip blank rows
                ws[f'A{row}'] = stat_name
                ws[f'C{row}'] = stat_value
                if format_str:
                    ws[f'C{row}'].number_format = format_str
                ws[f'C{row}'].font = Font(bold=True)
            row += 1
        
        # Scenario Comparison
        row += 2
        ws[f'A{row}'] = "SCENARIO COMPARISON"
        ws[f'A{row}'].font = self.subtitle_font
        ws.merge_cells(f'A{row}:F{row}')
        row += 2
        
        # Headers for scenario comparison
        headers = ["Scenario", "Mean Harm", "Reduction %", "Annual Impact", "Impact Reduction"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
        row += 1
        
        # Scenario data
        status_quo_mean = scenario_results['Status Quo']['stats']['Mean Harm']
        status_quo_impact = scenario_results['Status Quo']['stats']['Annual Industry Impact (Mean)']
        
        for scenario_name, scenario_data in scenario_results.items():
            ws[f'A{row}'] = scenario_name
            ws[f'B{row}'] = scenario_data['stats']['Mean Harm']
            ws[f'B{row}'].number_format = self.currency_format
            
            if scenario_name != 'Status Quo':
                reduction = (1 - scenario_data['stats']['Mean Harm'] / status_quo_mean) * 100
                ws[f'C{row}'] = reduction / 100
                ws[f'C{row}'].number_format = self.percent_format
                ws[f'C{row}'].fill = self.good_fill
            
            ws[f'D{row}'] = scenario_data['stats']['Annual Industry Impact (Mean)']
            ws[f'D{row}'].number_format = self.currency_format
            
            if scenario_name != 'Status Quo':
                impact_reduction = status_quo_impact - scenario_data['stats']['Annual Industry Impact (Mean)']
                ws[f'E{row}'] = impact_reduction
                ws[f'E{row}'].number_format = self.currency_format
                ws[f'E{row}'].fill = self.good_fill
            
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 15

    def create_detailed_results_sheet(self, results):
        """Create sheet with full simulation results"""
        ws = self.wb.create_sheet("Detailed Results")
        
        # Title
        ws['A1'] = "Full Simulation Results (First 1000 Records)"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:J1')
        
        # Add data
        row = 3
        
        # Headers
        headers = [
            "Simulation #", "Service Cost", "Hidden Fees", "Service Failure",
            "Service Failure Harm", "Damage Occurred", "Damage Value",
            "Claim Denied", "Damage Harm", "Total Harm"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.border
        
        # Data (first 1000 rows to keep file size manageable)
        for idx, row_data in results.head(1000).iterrows():
            row += 1
            ws.cell(row=row, column=1, value=idx + 1)
            ws.cell(row=row, column=2, value=row_data['service_cost']).number_format = self.currency_format
            ws.cell(row=row, column=3, value=row_data['hidden_fees']).number_format = self.currency_format
            ws.cell(row=row, column=4, value="Yes" if row_data['service_failure'] else "No")
            ws.cell(row=row, column=5, value=row_data['service_failure_harm']).number_format = self.currency_format
            ws.cell(row=row, column=6, value="Yes" if row_data['damage_occurred'] else "No")
            ws.cell(row=row, column=7, value=row_data['damage_value']).number_format = self.currency_format
            ws.cell(row=row, column=8, value="Yes" if row_data['claim_denied'] else "No")
            ws.cell(row=row, column=9, value=row_data['damage_harm']).number_format = self.currency_format
            ws.cell(row=row, column=10, value=row_data['total_harm']).number_format = self.currency_format
            
            # Apply conditional formatting for total harm
            if row_data['total_harm'] > 5000:
                ws.cell(row=row, column=10).fill = self.bad_fill
            elif row_data['total_harm'] > 1000:
                ws.cell(row=row, column=10).fill = self.neutral_fill
            else:
                ws.cell(row=row, column=10).fill = self.good_fill
        
        # Adjust column widths
        for col in range(1, 11):
            if col == 1:
                ws.column_dimensions[chr(64 + col)].width = 12
            else:
                ws.column_dimensions[chr(64 + col)].width = 18

    def create_percentile_analysis_sheet(self, results):
        """Create percentile analysis sheet"""
        ws = self.wb.create_sheet("Percentile Analysis")
        
        ws['A1'] = "Consumer Harm Percentile Analysis"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:E1')
        
        # Calculate percentiles
        percentiles = [1, 5, 10, 25, 50, 75, 80, 85, 90, 95, 99]
        
        row = 3
        # Headers
        headers = ["Percentile", "Harm Amount", "Cumulative Customers", "% of Customers", "Interpretation"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
        
        row += 1
        total_customers = len(results)
        
        for p in percentiles:
            harm_value = results['total_harm'].quantile(p/100)
            cumulative_customers = int(total_customers * p / 100)
            
            ws.cell(row=row, column=1, value=f"{p}th")
            ws.cell(row=row, column=2, value=harm_value).number_format = self.currency_format
            ws.cell(row=row, column=3, value=cumulative_customers).number_format = self.number_format
            ws.cell(row=row, column=4, value=p/100).number_format = self.percent_format
            
            # Interpretation
            if p <= 25:
                interpretation = "Low harm"
                fill = self.good_fill
            elif p <= 75:
                interpretation = "Moderate harm"
                fill = self.neutral_fill
            else:
                interpretation = "High to extreme harm"
                fill = self.bad_fill
            
            ws.cell(row=row, column=5, value=interpretation)
            for col in range(1, 6):
                ws.cell(row=row, column=col).fill = fill
            
            row += 1
        
        # Add summary statistics
        row += 2
        ws[f'A{row}'] = "Distribution Insights"
        ws[f'A{row}'].font = self.section_font
        row += 1
        
        insights = [
            f"50% of customers experience harm ≤ ${results['total_harm'].quantile(0.5):,.0f}",
            f"25% of customers experience harm ≥ ${results['total_harm'].quantile(0.75):,.0f}",
            f"10% of customers experience harm ≥ ${results['total_harm'].quantile(0.9):,.0f}",
            f"5% of customers experience harm ≥ ${results['total_harm'].quantile(0.95):,.0f}",
            f"1% of customers experience harm ≥ ${results['total_harm'].quantile(0.99):,.0f}",
        ]
        
        for insight in insights:
            ws[f'A{row}'] = insight
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 20

    def create_harm_components_sheet(self, results):
        """Create harm components breakdown sheet"""
        ws = self.wb.create_sheet("Harm Components")
        
        ws['A1'] = "Consumer Harm Components Analysis"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:F1')
        
        # Component statistics
        components = {
            'Hidden Fees': results['hidden_fees'],
            'Service Failure Harm': results['service_failure_harm'],
            'Damage Harm (Denied Claims)': results['damage_harm']
        }
        
        row = 3
        # Headers
        headers = ["Component", "Mean", "Median", "Max", "% of Total", "Affected Customers"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
        
        row += 1
        total_mean_harm = results['total_harm'].mean()
        
        for component_name, component_data in components.items():
            ws.cell(row=row, column=1, value=component_name)
            ws.cell(row=row, column=2, value=component_data.mean()).number_format = self.currency_format
            ws.cell(row=row, column=3, value=component_data.median()).number_format = self.currency_format
            ws.cell(row=row, column=4, value=component_data.max()).number_format = self.currency_format
            ws.cell(row=row, column=5, value=component_data.mean() / total_mean_harm).number_format = self.percent_format
            ws.cell(row=row, column=6, value=(component_data > 0).sum()).number_format = self.number_format
            row += 1
        
        # Correlation matrix
        row += 2
        ws[f'A{row}'] = "Component Correlation Matrix"
        ws[f'A{row}'].font = self.section_font
        row += 2
        
        corr_data = results[['service_cost', 'hidden_fees', 'service_failure_harm', 
                           'damage_harm', 'total_harm']].corr()
        
        # Headers
        corr_headers = [''] + list(corr_data.columns)
        for col, header in enumerate(corr_headers, 1):
            if col > 1:
                header = header.replace('_', ' ').title()
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = self.header_fill if col > 1 else None
        
        # Correlation values
        for idx, row_label in enumerate(corr_data.index):
            row += 1
            ws.cell(row=row, column=1, value=row_label.replace('_', ' ').title())
            ws.cell(row=row, column=1).font = Font(bold=True)
            ws.cell(row=row, column=1).fill = self.header_fill
            
            for col_idx, value in enumerate(corr_data.iloc[idx], 2):
                cell = ws.cell(row=row, column=col_idx, value=value)
                cell.number_format = '0.00'
                
                # Color scale for correlation
                if value > 0.7:
                    cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                elif value > 0.3:
                    cell.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
                elif value < -0.3:
                    cell.fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        for col in range(2, 7):
            ws.column_dimensions[chr(64 + col)].width = 18

    def create_scenario_comparison_sheet(self, scenario_results):
        """Create detailed scenario comparison sheet"""
        ws = self.wb.create_sheet("Scenario Comparison")
        
        ws['A1'] = "Reform Scenario Impact Analysis"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:G1')
        
        row = 3
        
        # Create comparison table
        scenarios = list(scenario_results.keys())
        metrics = [
            'Mean Harm', 'Median Harm', '90th Percentile', '95th Percentile',
            '99th Percentile', 'Customers with Zero Harm', 
            'Customers with Harm > $1000', 'Customers with Harm > $5000',
            'Annual Industry Impact (Mean)'
        ]
        
        # Headers
        headers = ['Metric'] + scenarios
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
        
        row += 1
        
        # Data
        for metric in metrics:
            ws.cell(row=row, column=1, value=metric)
            ws.cell(row=row, column=1).font = Font(bold=True)
            
            for col, scenario in enumerate(scenarios, 2):
                value = scenario_results[scenario]['stats'][metric]
                cell = ws.cell(row=row, column=col, value=value)
                
                if '$' in metric or 'Harm' in metric or 'Impact' in metric:
                    cell.number_format = self.currency_format
                else:
                    cell.number_format = self.number_format
                
                # Color coding for improvements
                if col > 2 and scenario != 'Status Quo':
                    status_quo_value = scenario_results['Status Quo']['stats'][metric]
                    if 'Customers with' in metric or 'Harm' in metric:
                        if value < status_quo_value:
                            cell.fill = self.good_fill
            row += 1
        
        # Cost-Benefit Analysis
        row += 2
        ws[f'A{row}'] = "Cost-Benefit Analysis"
        ws[f'A{row}'].font = self.section_font
        row += 2
        
        # Implementation costs (example values)
        implementation_costs = {
            'Status Quo': 0,
            'Moderate Reform': 191_000_000,  # From your analysis
            'Strong Reform': 297_000_000
        }
        
        headers = ['Scenario', 'Annual Consumer Benefit', 'Implementation Cost', 'Net Benefit', 'ROI']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
        
        row += 1
        status_quo_impact = scenario_results['Status Quo']['stats']['Annual Industry Impact (Mean)']
        
        for scenario in scenarios:
            ws.cell(row=row, column=1, value=scenario)
            
            current_impact = scenario_results[scenario]['stats']['Annual Industry Impact (Mean)']
            benefit = status_quo_impact - current_impact if scenario != 'Status Quo' else 0
            cost = implementation_costs[scenario]
            net_benefit = benefit - cost
            roi = (benefit / cost - 1) if cost > 0 else 0
            
            ws.cell(row=row, column=2, value=benefit).number_format = self.currency_format
            ws.cell(row=row, column=3, value=cost).number_format = self.currency_format
            ws.cell(row=row, column=4, value=net_benefit).number_format = self.currency_format
            ws.cell(row=row, column=5, value=roi).number_format = self.percent_format
            
            if net_benefit > 0:
                for col in range(2, 6):
                    ws.cell(row=row, column=col).fill = self.good_fill
            
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        for col in range(2, 8):
            ws.column_dimensions[chr(64 + col)].width = 20

    def create_charts_sheet(self, results, scenario_results):
        """Create sheet with charts"""
        ws = self.wb.create_sheet("Charts")
        
        ws['A1'] = "Visual Analysis"
        ws['A1'].font = self.title_font
        
        # Prepare data for charts
        # 1. Harm Distribution Data (binned for chart)
        bins = [0, 100, 500, 1000, 2000, 5000, 10000, float('inf')]
        labels = ['$0-100', '$100-500', '$500-1k', '$1k-2k', '$2k-5k', '$5k-10k', '>$10k']
        harm_dist = pd.cut(results['total_harm'], bins=bins, labels=labels).value_counts().sort_index()
        
        # Write distribution data
        ws['A3'] = "Harm Distribution"
        ws['A3'].font = self.section_font
        ws['A4'] = "Range"
        ws['B4'] = "Count"
        
        row = 5
        for label, count in harm_dist.items():
            ws[f'A{row}'] = str(label)
            ws[f'B{row}'] = int(count)
            row += 1
        
        # Create bar chart for distribution
        chart1 = BarChart()
        chart1.type = "col"
        chart1.style = 10
        chart1.title = "Consumer Harm Distribution"
        chart1.y_axis.title = 'Number of Customers'
        chart1.x_axis.title = 'Harm Amount Range'
        
        data = Reference(ws, min_col=2, min_row=4, max_row=row-1, max_col=2)
        cats = Reference(ws, min_col=1, min_row=5, max_row=row-1)
        chart1.add_data(data, titles_from_data=True)
        chart1.set_categories(cats)
        chart1.shape = 4
        ws.add_chart(chart1, "D3")
        
        # 2. Scenario Comparison Chart
        ws['A15'] = "Scenario Comparison"
        ws['A15'].font = self.section_font
        ws['A16'] = "Scenario"
        ws['B16'] = "Mean Harm"
        ws['C16'] = "Annual Impact (Millions)"
        
        row = 17
        for scenario, data in scenario_results.items():
            ws[f'A{row}'] = scenario
            ws[f'B{row}'] = data['stats']['Mean Harm']
            ws[f'C{row}'] = data['stats']['Annual Industry Impact (Mean)'] / 1_000_000
            row += 1
        
        # Create comparison chart
        chart2 = BarChart()
        chart2.type = "col"
        chart2.style = 11
        chart2.title = "Reform Scenario Comparison"
        chart2.y_axis.title = 'Mean Harm ($)'
        chart2.x_axis.title = 'Scenario'
        
        data = Reference(ws, min_col=2, min_row=16, max_row=row-1, max_col=2)
        cats = Reference(ws, min_col=1, min_row=17, max_row=row-1)
        chart2.add_data(data, titles_from_data=True)
        chart2.set_categories(cats)
        ws.add_chart(chart2, "D15")
        
        # 3. Component Breakdown Pie Chart
        ws['A27'] = "Harm Components"
        ws['A27'].font = self.section_font
        ws['A28'] = "Component"
        ws['B28'] = "Average Amount"
        
        components = [
            ('Hidden Fees', results['hidden_fees'].mean()),
            ('Service Failures', results['service_failure_harm'].mean()),
            ('Denied Claims', results['damage_harm'].mean())
        ]
        
        row = 29
        for comp_name, comp_value in components:
            ws[f'A{row}'] = comp_name
            ws[f'B{row}'] = comp_value
            row += 1
        
        # Create pie chart
        pie = PieChart()
        labels = Reference(ws, min_col=1, min_row=29, max_row=31)
        data = Reference(ws, min_col=2, min_row=28, max_row=31)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        pie.title = "Average Harm by Component"
        ws.add_chart(pie, "D27")

    def create_parameters_sheet(self):
        """Create sheet documenting simulation parameters"""
        ws = self.wb.create_sheet("Parameters")
        
        ws['A1'] = "Monte Carlo Simulation Parameters"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:D1')
        
        ws['A3'] = "Simulation Configuration"
        ws['A3'].font = self.section_font
        
        row = 4
        configs = [
            ("Number of Simulations", N_SIMULATIONS, self.number_format),
            ("Annual Transactions", ANNUAL_TRANSACTIONS, self.number_format),
            ("Service Failure Penalty", 1000, self.currency_format),
            ("Random Seed", 42, None)
        ]
        
        for config_name, config_value, format_str in configs:
            ws[f'A{row}'] = config_name
            ws[f'C{row}'] = config_value
            if format_str:
                ws[f'C{row}'].number_format = format_str
            row += 1
        
        row += 1
        ws[f'A{row}'] = "Distribution Parameters (Triangular)"
        ws[f'A{row}'].font = self.section_font
        row += 1
        
        # Headers
        headers = ["Parameter", "Minimum", "Mode (Most Likely)", "Maximum"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
        
        row += 1
        
        # Parameters
        for param_name, param_values in PARAMS.items():
            ws[f'A{row}'] = param_name.replace('_', ' ').title()
            ws[f'B{row}'] = param_values['min']
            ws[f'C{row}'] = param_values['mode']
            ws[f'D{row}'] = param_values['max']
            
            # Format based on parameter type
            if 'cost' in param_name or 'fee' in param_name or 'value' in param_name:
                for col in range(2, 5):
                    ws.cell(row=row, column=col).number_format = self.currency_format
            elif 'prob' in param_name or 'rate' in param_name:
                for col in range(2, 5):
                    ws.cell(row=row, column=col).number_format = self.percent_format
            
            row += 1
        
        # Add notes
        row += 2
        ws[f'A{row}'] = "Notes:"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        notes = [
            "• Triangular distributions used to model most likely scenarios with uncertainty",
            "• Parameters based on analysis of 2,500+ customer reviews and legal documents",
            "• Service failure penalty represents additional costs from delays and disruptions",
            "• Random seed ensures reproducibility of results"
        ]
        
        for note in notes:
            ws[f'A{row}'] = note
            ws.merge_cells(f'A{row}:D{row}')
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15

    def save_workbook(self):
        """Save the completed workbook"""
        self.wb.save(self.filename)
        print(f"Excel file saved as: {self.filename}")

def main():
    """Main execution function"""
    print("Generating Excel report from Monte Carlo simulation...")
    print("=" * 50)
    
    # Run simulations
    print("\nRunning base simulation...")
    results = run_monte_carlo_simulation()
    stats = calculate_statistics(results)
    
    print("Running scenario analysis...")
    scenarios = {
        'Status Quo': PARAMS,
        'Moderate Reform': {
            'base_service_cost': PARAMS['base_service_cost'],
            'hidden_fees': {'min': 0, 'mode': 150, 'max': 500},
            'service_failure_prob': {'min': 0.10, 'mode': 0.20, 'max': 0.30},
            'claim_denial_prob': {'min': 0.40, 'mode': 0.60, 'max': 0.80},
            'damage_occurrence_rate': PARAMS['damage_occurrence_rate'],
            'average_damage_value': PARAMS['average_damage_value']
        },
        'Strong Reform': {
            'base_service_cost': PARAMS['base_service_cost'],
            'hidden_fees': {'min': 0, 'mode': 50, 'max': 200},
            'service_failure_prob': {'min': 0.05, 'mode': 0.10, 'max': 0.15},
            'claim_denial_prob': {'min': 0.20, 'mode': 0.35, 'max': 0.50},
            'damage_occurrence_rate': {'min': 0.03, 'mode': 0.08, 'max': 0.15},
            'average_damage_value': PARAMS['average_damage_value']
        }
    }
    
    scenario_results = {}
    for scenario_name, scenario_params in scenarios.items():
        scenario_res = run_monte_carlo_simulation(scenario_params, n_sims=N_SIMULATIONS)
        scenario_stats = calculate_statistics(scenario_res)
        scenario_results[scenario_name] = {
            'results': scenario_res,
            'stats': scenario_stats
        }
    
    # Create Excel file
    print("\nCreating Excel workbook...")
    exporter = ExcelExporter()
    
    print("- Adding Executive Summary...")
    exporter.create_summary_sheet(stats, scenario_results)
    
    print("- Adding Detailed Results...")
    exporter.create_detailed_results_sheet(results)
    
    print("- Adding Percentile Analysis...")
    exporter.create_percentile_analysis_sheet(results)
    
    print("- Adding Harm Components...")
    exporter.create_harm_components_sheet(results)
    
    print("- Adding Scenario Comparison...")
    exporter.create_scenario_comparison_sheet(scenario_results)
    
    print("- Adding Charts...")
    exporter.create_charts_sheet(results, scenario_results)
    
    print("- Adding Parameters Documentation...")
    exporter.create_parameters_sheet()
    
    print("\nSaving workbook...")
    exporter.save_workbook()
    
    # Also save raw data as CSV for additional analysis
    results.to_csv('monte_carlo_raw_data.csv', index=False)
    print("Raw data also saved as: monte_carlo_raw_data.csv")
    
    print("\n✅ Excel report generation complete!")
    print("\nThe Excel file contains:")
    print("  1. Executive Summary - Key findings and scenario comparison")
    print("  2. Detailed Results - First 1,000 simulation records")
    print("  3. Percentile Analysis - Distribution breakdown")
    print("  4. Harm Components - Analysis by harm type")
    print("  5. Scenario Comparison - Reform impact analysis")
    print("  6. Charts - Visual representations")
    print("  7. Parameters - Complete documentation")

if __name__ == "__main__":
    main()